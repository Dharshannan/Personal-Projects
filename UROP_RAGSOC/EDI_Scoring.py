from openpyxl import Workbook, load_workbook
import numpy as np

# =============================================================================
# We will 1st read data from the Excel file

filename = str(input('Data file name: ')) + '.xlsx'
wb = load_workbook(filename)
sh = wb['Sheet1']

row_ct = sh.max_row
col_ct =sh.max_column

# 1st we append the EDI
list1 = []

for i in range(2,row_ct + 1):
    list1.append(str(sh.cell(row=i, column=1).value))

#print(list1)

# We append all the different group/variable in the list
res = []
[res.append(x) for x in list1 if x not in res]

# Count the number of times each of these group/variable repeats
scores = []
a = 0
while a < len(res):
    count = 0
    for i in range(0, len(list1)):
        if list1[i] == res[a]:
            count += 1
    scores.append(count)
    a += 1
    
# Get only the score values (Non-Repeating)
scores2 = []
[scores2.append(x) for x in scores if x not in scores2]

# Sort the scores in ascending and descending order
scores2.sort()
asc = scores2.copy()
scores2.sort(reverse=True)
dsc = scores2.copy()

# Now index the scores (scores, index) this will point minority to highest score and majority to lowest
scoreindex = []
for i in range(0, len(scores2)):
    scoreindex.append((dsc[i],i))
    
# Now we assign the scores based of the index of the asc (ascending order list)
presort = list(zip(res,scores))
#print(presort)
for i in range(0, len(presort)):
    presort[i] = list(presort[i]) # Turn tuple to list
    
for i in range(0, len(presort)):
    index = asc.index(presort[i][1])
    for j in range(0, len(scoreindex)):
        if scoreindex[j][1] == index:
            presort[i][1] = scoreindex[j][0]
            
#print(presort)

# Now we assign these values to the original list
ediscores = []
for i in range(0, len(list1)):
    for j in range(0, len(presort)):
        if list1[i] == presort[j][0]:
            ediscores.append((list1[i],presort[j][1]))

print(ediscores)    

# We write this into the 2nd column of the document

for i in range(2,row_ct + 1):
    sh.cell(row=i, column=2).value = ediscores[i-2][1]
    

wb.save(filename)