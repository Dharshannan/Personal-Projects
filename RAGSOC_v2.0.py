from openpyxl import Workbook, load_workbook
import numpy as np

# =============================================================================
# We will 1st read data from the Excel file

filename = str(input('Data file name: ')) + '.xlsx'
wb = load_workbook(filename)
sh = wb['Sheet1']

row_ct = sh.max_row
col_ct =sh.max_column

# 1st we append the names and emails
Names = []
Email = []

for i in range(2,row_ct + 1):
    Names.append(str(sh.cell(row=i, column=1).value))
    Email.append(str(sh.cell(row=i, column=2).value))
    
# Next we append the attributes with while loop
count = 3 # We start from the 3rd row, this is always the case
ATTs = [] # Now we create a list for all the attributes
for i in range(0, (col_ct - 2)):
    ATTs.append([]) # We append empty lists
    
# We start appending the attributes
while count <= col_ct:
    for i in range(2,row_ct + 1):
        ATTs[count - 3].append(sh.cell(row=i, column=count).value)
    count += 1

list1 = list(zip(Names, Email))
for i in range(0, len(ATTs)):
    for j in range(0, len(list1)):
        list1[j] = list1[j] + (ATTs[i][j],)

# =============================================================================
# Now we implement the weightages ;) 
count2 = 3
weights = []
while count2 <= (col_ct - 1):
    weight = float(input(f'ATT{count2 - 2} Weightage(%): '))/100
    weights.append(weight)
    count2 += 1
    
# Check and verify if the weights add up to 100%
percentsum = round(sum(weights)*100)
if (percentsum < 100) or (percentsum > 100):
    print(f'Caution the sum of weights is not exactly 100%, the weights add up to: {percentsum}%')
    print('You may wish to proceed if the sum is very close/approximately 100%')
# =============================================================================
# Now we update the list with weighted means

for i in range(0, len(list1)):
    weight_sum = 0
    for j in range(0, len(weights)):
        weight_sum += weights[j]*list1[i][j+2]/max(ATTs[j]) # We normalize by dividing the score with max score
    list1[i] = list1[i] + (weight_sum,)
    
# We will swap the last 2 atts positions in the list to make it so we can copy paste the older code lol :)

for i in range(0, len(list1)):
    list1[i] = list(list1[i]) # Tuples are immutable hence convert to list
    temp = list1[i][-1]
    list1[i][-1] = list1[i][-2] # Swap positions
    list1[i][-2] = temp
    list1[i] = tuple(list1[i]) # Convert back to tuple

# =============================================================================
# Now we re-use old RAGSOC Code lul
import statistics as st
import random as rn

totalstud = len(list1)
totalgrp = int(input('Total No of Groups: '))
minstudpg = int(input('Minimum No of Students per Group: '))

Groups0 = []
for ll in range(0,totalgrp):
    shitt = []
    for ff in range(0,minstudpg):
        poopp = []
        shitt.append(poopp)
    Groups0.append(shitt)

abc = 0
liss = []
liss1= []
while abc < totalgrp*minstudpg:
    tt = rn.randint(1,totalgrp) - 1
    ss = rn.randint(1,minstudpg) - 1
    vv = rn.randint(1,totalstud) - 1
    if [tt,ss] not in liss and vv not in liss1:
       liss.append([tt,ss])
       liss1.append(vv)
       Groups0[tt][ss] = list1[vv]
       abc += 1
         
testwmean = []
testatt4 =[]

for i in range(0,totalgrp):
    aa = 0
    sumgroup1a = 0
    sumgroup2a = 0
    while aa < minstudpg:
        sumgroup1a += Groups0[i][aa][-2]
        sumgroup2a += Groups0[i][aa][-1]
        aa += 1
        
    testwmean.append(sumgroup1a)
    testatt4.append(sumgroup2a)

    
meanwmean = st.mean(testwmean)
meanATT4 = st.mean(testatt4)
stdwmean = st.pstdev(testwmean)
stdATT4 = st.pstdev(testatt4)


Groups = []
for l in range(0,totalgrp):
    shit = []
    for f in range(0,minstudpg):
        poop = []
        shit.append(poop)
    Groups.append(shit)

combcount = 0 # To check number of groups combinations tried
while True:
    combcount += 1
    restud = []
    a = 0
    lis = []
    lis1 = []
    while a < len(list1):
         t = rn.randint(1,totalgrp) - 1
         s = rn.randint(1,minstudpg) - 1
         v = rn.randint(1,len(list1)) - 1
         if [t,s] not in lis and v not in lis1 and a < totalgrp*minstudpg:
            lis.append([t,s])
            lis1.append(v)
            Groups[t][s] = list1[v]
            a += 1
         elif v not in lis1 and a >= totalgrp*minstudpg:
            lis1.append(v)
            restud = restud + [list1[v]]
            a += 1
       
    avgwmeang = []
    avgatt4 =[]
       
    for i in range(0,totalgrp):
       a = 0
       sumgroup1 = 0
       sumgroup2 = 0
       
       while a < minstudpg:
           sumgroup1 += Groups[i][a][-2]
           sumgroup2 += Groups[i][a][-1]
           a += 1
       #print(sumgroup1 , sumgroup2)
       avgwmeang.append(sumgroup1)
       avgatt4.append(sumgroup2)
       
       
    fml = []
    for item in avgwmeang:
       #print(item)
       if item >= (meanwmean - stdwmean) and item <= (meanwmean + stdwmean):
           #print('Yes')
           fml.append('Yes')
       else:
           #print('No')
           fml.append('No')
    #print(fml)
    
    
    imo = []
    for itemss in avgatt4:
        if itemss >= round(meanATT4 - stdATT4) and itemss <= round(meanATT4 + stdATT4):
            imo.append('Yes')
        else:
            imo.append('No')
            
       
    shark = []
    for k in range(0,totalgrp):
       fish = 'Yes'
       shark.append(fish)
    #print(shark)
    if shark == fml and shark == imo:
        break
 
for n in range(0,len(restud)):
    Groups[n] = Groups[n] + [restud[n]]

print(Groups)
# =============================================================================

# Now we will write this into an Excel file

import xlsxwriter

new_file = str(input('Groups file name: ')) + '.xlsx'
outWorkbook = xlsxwriter.Workbook(new_file)
outSheet = outWorkbook.add_worksheet()
bold = outWorkbook.add_format({'bold':True})
bold.set_align('center')

outSheet.write('A1','Group',bold)
outSheet.write('B1','Names',bold)
outSheet.write('C1','Email',bold)

# Now we will write the attribute titles using a while loop
count3 = 3
ch = 'D' # We start with column D, this is also always the case
while count3 <= (col_ct + 1): # +1 to include the weighted means
    outSheet.write(f'{ch}1',f'ATT{count3 - 2}',bold)
    ch = str(chr(ord(ch) + 1))
    count3 += 1
    
# Now we will format and write the data in an organised fashion

# 1st we will do it for the groups, names and email (This is easy)
groups = []
names = []
email = []

for itemsss in range(0,len(Groups)):
    for crap in range(0,len(Groups[itemsss])):
        groups.append(itemsss+1)
        names.append(Groups[itemsss][crap][0])
        email.append(Groups[itemsss][crap][1])

# Next we do it for the attributes (This is a little complex)
atts = []
for i in range(0, (col_ct - 1)):  # Changed -2 to -1 to account for the extra weighted means
    atts.append([]) # We append empty lists

count4 = 2
while count4 <= col_ct:
    for x in range(0, len(Groups)):
        for y in range(0, len(Groups[x])):
            atts[count4 - 2].append(Groups[x][y][count4])
    count4 += 1
    
# Now we write the data into our workbook
for k in range(0,len(groups)):
    outSheet.write(k+1,0,groups[k])
    outSheet.write(k+1,1,names[k])
    outSheet.write(k+1,2,email[k])
    
count5 = 3
while count5 <= (col_ct + 1):
    for m in range(0, len(groups)):
        outSheet.write(m+1,count5,atts[count5 - 3][m])
    count5 += 1
    
outWorkbook.close()
# =============================================================================
# Now lastly, we print the confirmation
print(f'Done, number of groups combinations tried: {combcount}')
# =============================================================================