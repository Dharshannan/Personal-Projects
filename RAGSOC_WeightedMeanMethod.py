from openpyxl import Workbook, load_workbook
import numpy as np

wb = load_workbook('FCTTDataSpreadsheet.xlsx')
sh = wb['Sheet1']

row_ct = sh.max_row
col_ct =sh.max_column

Names = []
Email = []
ATT1 = []
ATT2 = []
ATT3 = []
ATT4 = []

for i in range(2,row_ct + 1):
    Names.append(str(sh.cell(row=i, column=1).value))
    Email.append(str(sh.cell(row=i, column=2).value))
    ATT1.append(sh.cell(row=i, column=3).value)
    ATT2.append(sh.cell(row=i, column=4).value)
    ATT3.append(sh.cell(row=i, column=5).value)
    ATT4.append(sh.cell(row=i, column=6).value)
    
def normalizeatt(att):
    maxatt = np.max(att)
    normalizedlist = []
    for i in range(0, len(att)):
        normalizedlist.append(att[i]/maxatt)
    return(normalizedlist)
    

normalizeatt1 = normalizeatt(ATT1)
normalizeatt2 = normalizeatt(ATT2)
normalizeatt3 = normalizeatt(ATT3)

# weighted mean of 3 attr

weighted_mean = []
weight1 = int(input('Top Priority Weight (%): '))/100
weight2 = int(input('2nd Priority Weight (%): '))/100
weight3 = int(input('3rd Priority Weight (%): '))/100
for i in range(0, len(ATT1)):
    weighted_sum = normalizeatt1[i]*weight1 + normalizeatt2[i]*weight2 + normalizeatt3[i]*weight3
    weighted_mean.append(weighted_sum)
    

list0 = zip(Names,Email,ATT1,ATT2,ATT3,weighted_mean,ATT4)
list1 = list(list0)


import statistics as st
import random as rn

totalstud = len(list1)
totalgrp = int(input('Total No of Groups: '))
minstudpg = int(input('Minimum No of Students per Group: '))

Groups0 = []
for ll in range(0,totalgrp):
    shitt = []
    for ff in range(0,minstudpg):
        poopp = []
        shitt.append(poopp)
    Groups0.append(shitt)

abc = 0
liss = []
liss1= []
while abc < totalgrp*minstudpg:
      tt = rn.randint(1,totalgrp) - 1
      ss = rn.randint(1,minstudpg) - 1
      vv = rn.randint(1,totalstud) - 1
      if [tt,ss] not in liss and vv not in liss1:
         liss.append([tt,ss])
         liss1.append(vv)
         Groups0[tt][ss] = list1[vv]
         abc += 1
         
testwmean = []
testatt4 =[]

for i in range(0,totalgrp):
    aa = 0
    sumgroup1a = 0
    sumgroup2a = 0
    while aa < minstudpg:
        sumgroup1a += Groups0[i][aa][5]
        sumgroup2a += Groups0[i][aa][6]
        aa += 1
        
    testwmean.append(sumgroup1a)
    testatt4.append(sumgroup2a)

    
meanwmean = st.mean(testwmean)
meanATT4 = st.mean(testatt4)
stdwmean = st.pstdev(testwmean)
stdATT4 = st.pstdev(testatt4)


Groups = []
for l in range(0,totalgrp):
    shit = []
    for f in range(0,minstudpg):
        poop = []
        shit.append(poop)
    Groups.append(shit)


while True:
 restud = []
 a = 0
 lis = []
 lis1 = []
 while a < len(list1):
      t = rn.randint(1,totalgrp) - 1
      s = rn.randint(1,minstudpg) - 1
      v = rn.randint(1,len(list1)) - 1
      if [t,s] not in lis and v not in lis1 and a < totalgrp*minstudpg:
         lis.append([t,s])
         lis1.append(v)
         Groups[t][s] = list1[v]
         a += 1
      elif v not in lis1 and a >= totalgrp*minstudpg:
         lis1.append(v)
         restud = restud + [list1[v]]
         a += 1

 avgwmeang = []
 avgatt4 =[]

 for i in range(0,totalgrp):
    a = 0
    sumgroup1 = 0
    sumgroup2 = 0
    
    while a < minstudpg:
        sumgroup1 += Groups[i][a][5]
        sumgroup2 += Groups[i][a][6]
        a += 1
    #print(sumgroup1 , sumgroup2)
    avgwmeang.append(sumgroup1)
    avgatt4.append(sumgroup2)


 fml = []
 for item in avgwmeang:
    #print(item)
    if item >= (meanwmean - stdwmean) and item <= (meanwmean + stdwmean):
        #print('Yes')
        fml.append('Yes')
    else:
        #print('No')
        fml.append('No')
 #print(fml)
 
 
 imo = []
 for itemss in avgatt4:
     if itemss >= round(meanATT4 - stdATT4) and itemss <= round(meanATT4 + stdATT4):
         imo.append('Yes')
     else:
         imo.append('No')
         

 shark = []
 for k in range(0,totalgrp):
    fish = 'Yes'
    shark.append(fish)
 #print(shark)
 if shark == fml and shark == imo:
     break
 
for n in range(0,len(restud)):
    Groups[n] = Groups[n] + [restud[n]]


print(Groups)

import xlsxwriter

outWorkbook = xlsxwriter.Workbook('FCTT_GROUPS.xlsx')
outSheet = outWorkbook.add_worksheet()
bold = outWorkbook.add_format({'bold':True})
bold.set_align('center')

outSheet.write('A1','Group',bold)
outSheet.write('B1','Names',bold)
outSheet.write('C1','Email',bold)
outSheet.write('D1','ATT1',bold)
outSheet.write('E1','ATT2',bold)
outSheet.write('F1','ATT3',bold)
outSheet.write('G1','ATT4',bold)

groups = []
names = []
email = []
att1 = []
att2 = []
att3 = []
att4 = []
for itemsss in range(0,len(Groups)):
    for crap in range(0,len(Groups[itemsss])):
        groups.append(itemsss+1)
        names.append(Groups[itemsss][crap][0])
        email.append(Groups[itemsss][crap][1])
        att1.append(Groups[itemsss][crap][2])
        att2.append(Groups[itemsss][crap][3])
        att3.append(Groups[itemsss][crap][4])
        att4.append(Groups[itemsss][crap][6])

for k in range(0,len(groups)):
    outSheet.write(k+1,0,groups[k])

for v in range(0,len(names)):
    outSheet.write(v+1,1,names[v])

for s in range(0,len(email)):
    outSheet.write(s+1,2,email[s])
    
for q in range(0,len(att1)):
    outSheet.write(q+1,3,att1[q])
    
for t in range(0,len(att2)):
    outSheet.write(t+1,4,att2[t])
    
for o in range(0,len(att3)):
    outSheet.write(o+1,5,att3[o])

for z in range(0,len(att4)):
    outSheet.write(z+1,6,att4[z])
 
outWorkbook.close()

